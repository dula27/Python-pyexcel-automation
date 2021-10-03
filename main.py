import openpyxl
import os

source_files = '/mnt/c/Users/Abdullah/Desktop/cd/github/pyexcel/2021/'
excel_files = os.listdir('/mnt/c/Users/Abdullah/Desktop/cd/github/pyexcel/2021/')
source = []

dest_file = '/mnt/c/Users/Abdullah/Desktop/cd/github/pyexcel/source_2021/master_file.xlsx'
dwb = openpyxl.Workbook()
dwb.save(dest_file)
wb = openpyxl.load_workbook(dest_file)
ws = wb.active

ws.cell(row = 1, column = 1).value = 'StudentName'
ws.cell(row = 1, column = 2).value = 'Date'
ws.cell(row = 1, column = 3).value = 'To Student/Envelope'
ws.cell(row = 1, column = 4).value = 'To Student Account'
ws.cell(row = 1, column = 5).value = 'Earned - CARES'

for excel in excel_files:
    source.append(source_files + excel)

for file in source:
    print(file)
    date = file[-13:-5]
    
    workbook = openpyxl.load_workbook(file,data_only=True)
    worksheet = workbook['Values']
    
    csize = 1
    while (worksheet.cell(row = csize, column = 1).value != None):
        csize += 1
        
    print(csize)

    for i in range (2, csize):
        # reading cell value from source excel file
        A = worksheet.cell(row = i, column = 1)
        # writing the read value to destination excel file
        # ws.cell(row = i, column = 1).value = A.value

        # writing the read value to destination excel file
        # ws.cell(row = i, column = 2).value = int(date)

        # reading cell value from source excel file
        M = worksheet.cell(row = i, column = 13)
        # writing the read value to destination excel file
        # ws.cell(row = i, column = 3).value = int(M.value)

        # reading cell value from source excel file
        N = worksheet.cell(row = i, column = 14)
        # writing the read value to destination excel file
        # ws.cell(row = i, column = 4).value = int(N.value)

        # writing the read value to destination excel file
        # ws.cell(row = i, column = 5).value = (int(N.value) + int(M.value))
        if(M.value == None):
            M.value = 0
        if(N.value == None):
            N.value = 0
            
        P = int(N.value) + int(M.value)
        ws.append([A.value, int(date), int(M.value), int(N.value), P])
    
    print(csize)

    wb.save(str(dest_file))